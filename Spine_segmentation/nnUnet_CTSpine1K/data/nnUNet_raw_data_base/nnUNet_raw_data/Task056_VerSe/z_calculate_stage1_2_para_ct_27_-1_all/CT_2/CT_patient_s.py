###############################################################################################################
#计算一张图片的密度
import os 
import numpy as np
import SimpleITK as sitk
from collections import Counter
import json

from tqdm import tqdm

import pandas as pd
import openpyxl

#分割参数
K_seg = 27
C_seg = -1
k_dili = 5
iter_dili = 6

# spine_id_l = [18,19,20,21]
spine_id_l = [18,19,20,21]

file_list = ["result_november","result_stage2"]

ct_p = "./CT_patient_result_cancellous_s"
# target_file_name = "view_all_spine.nii.gz"
target_file_name = "view_all_cancellous.nii.gz"

if not os.path.exists(ct_p):
    os.makedirs(ct_p)

#####################
xml_pth = "./193layer3.xlsx"
# xml_pth = "./all.xlsx"
wb = openpyxl.load_workbook(xml_pth)
# sheet = wb.active
sheet_list = wb.get_sheet_names()
print(sheet_list)

def write_to_excel(path,sheet_str,info,data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data.insert(0,list(info))
    sheet.title = sheet_str

    for row_index, row_item in enumerate(data):

        for col_index, col_item in enumerate(row_item):
            sheet.cell(row=row_index+1,column= col_index+1,value=col_item)
    workbook.save(path)

for sheet_name in sheet_list:             
    sheet = wb[sheet_name]
    print("sheet: {}".format(sheet_name))

    if(len(sheet_name)>4):
        file_list_n = "result_november"
        print("process result_november ")
    else:
        file_list_n = "result_stage2"
        print("process result_stage2 ")
#    #标注路径
#     label_root_pth = "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/label/".format(file_list_n)
#     #原图路径
#     img_root_pth =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/img_o/".format(file_list_n)
#     #分割结果
#     seg_root_pth =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/seg/".format(file_list_n)

#     out_root_pt h =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/out_ct_begin_27_-1_20231228_2/".format(file_list_n)


    spine_seg_root = "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/seg".format(file_list_n)
    cortex_seg_root = "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/out_ct_begin_27_-1_20231228_2".format(file_list_n)    ##seg,皮质，松质
    img_or_root = "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/img_o".format(file_list_n)
    
    name_list = []
    id_all = []
    id1 = []
    id2 = []
    id3 = []
    # id4 = []

    for cell in sheet['A']:
        name_list.append(cell.value)
    for cell in sheet['B']:
        id1.append(cell.value)
    id_all.append(id1)
    for cell in sheet['C']:
        id2.append(cell.value)
    id_all.append(id2)
    for cell in sheet['D']:
        id3.append(cell.value)
    id_all.append(id3)
    # for cell in sheet['E']:
    #     id4.append(cell.value)
    print(name_list)
    Gn = sheet_name
    for id,PAn in tqdm(enumerate(name_list)):
        file = "predict_{}_{}".format(Gn,PAn)
    
    # seg_pred_file_root = os.listdir(spine_seg_root)





    # for file in tqdm(seg_pred_file_root):
    #     Gn = file.split("_")[1]
    #     PAn = file.split("_")[2]
        spine_seg_pth = os.path.join(spine_seg_root,file,PAn+".nii.gz")
        img_or_path = os.path.join(img_or_root,"imagesTs_"+Gn+"_"+PAn,PAn+"_0000.nii.gz")
        cortex_seg_path = os.path.join(cortex_seg_root,file,"{}{}_K{}C{}k{}i{}".format(Gn,PAn,str(K_seg),str(C_seg),str(k_dili),str(iter_dili)),target_file_name)
        ct_f = os.path.join(ct_p,'{}.json'.format(file))
        ct_f2 = os.path.join(ct_p,'{}_py.json'.format(file))

        img = sitk.ReadImage(img_or_path)
        origin =img.GetOrigin()
        direction = img.GetDirection()
        space = img.GetSpacing()
        img = sitk.GetArrayFromImage(img)

        predict = sitk.ReadImage(spine_seg_pth)
        predict = sitk.GetArrayFromImage(predict)

        out = sitk.ReadImage(cortex_seg_path)
        out = sitk.GetArrayFromImage(out)

        c,h,w = predict.shape
        # slice_num = predict.shape[0]
        slice_num = 4
        ###################################
        spine_id_mask = np.zeros((c,h,w))
        s_mask = np.zeros((c,h,w))

        for s_id,spine_id in enumerate(spine_id_l):
            spine_id_mask[predict==spine_id]=1
        for s_id in range(3):
            z = id_all[s_id][id]
            s_mask[z,:,:]=1

            
        img_tmp = spine_id_mask*out*img
        img_tmp = img_tmp*s_mask
            

            
        intensity = img_tmp.sum()
        counter_this_slice = np.count_nonzero(img_tmp,axis=(0,1,2))
        density = intensity/counter_this_slice 
                
        file_name = file
        file_path = os.path.join(spine_seg_root,file)
        spine_id = spine_id
        json_list = {"file_name": file_name,"file_path": file_path,"spine_id": str(spine_id),"slice_num":int(slice_num),"density": float(density),"sum_of_intensity_in_this_id": float(intensity),"pixel_num_in_this_id": int(counter_this_slice)}
    ###json for read
        #     for item  in json_list:
        # #         print(json_list[item])
        #         with open(ct_f, 'a+', encoding='utf-8') as fp:
        #             line = json.dumps({item:json_list[item]},ensure_ascii=False)
        #             fp.write(line + '\n')
        #     with open(ct_f, 'a+', encoding='utf-8') as fp:
        #             line = json.dumps("--------------------------------------------------------",ensure_ascii=False)
        #             fp.write(line + '\n')
            
            ###json for python
        with open(ct_f2, 'a+', encoding='utf-8') as fp:
                    line = json.dumps(json_list,ensure_ascii=False)
                    fp.write(line + '\n')
            
print("end")
    
