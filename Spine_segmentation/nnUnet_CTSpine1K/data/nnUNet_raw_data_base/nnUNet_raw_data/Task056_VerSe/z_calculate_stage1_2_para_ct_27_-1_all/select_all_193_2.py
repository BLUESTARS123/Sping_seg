#实现选定部分切片和分割比较,绘制折线图

import os 
import numpy as np
import SimpleITK as sitk
from collections import Counter
import json

#绘图相关
import matplotlib.pyplot as plt

import pandas as pd
import openpyxl
from tqdm import tqdm


#松质id：
spine_id_l = [2,4,6,8]
seg_id_l = [18,19,20,21]

#分割参数
K_seg = 27
C_seg = -1
k_dili = 5
iter_dili = 6

#xml相关
xml_pth = "./193layer3.xlsx"
# xml_pth = "./all.xlsx"
wb = openpyxl.load_workbook(xml_pth)
# sheet = wb.active
sheet_list = wb.get_sheet_names()
print(sheet_list)






def write_to_excel(path,sheet_str,info,data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    data.insert(0,list(info))
    sheet.title = sheet_str

    for row_index, row_item in enumerate(data):

        for col_index, col_item in enumerate(row_item):
            sheet.cell(row=row_index+1,column= col_index+1,value=col_item)
    workbook.save(path)

#保存数据用于绘图
ans_1 = []
ans_2 = []

excel_list = []


# file_list_n = "result_stage2"


excel_path = "Sampling_vs_seg_193.xlsx"
sheet_str = "Spine"

info = ["file_name","Sampling","seg"]

for sheet_name in sheet_list:             
    sheet = wb[sheet_name]
    print("sheet: {}".format(sheet_name))

    if(len(sheet_name)>4):
        file_list_n = "result_november"
        print("process result_november ")
    else:
        file_list_n = "result_stage2"
        print("process result_stage2 ")
    #标注路径
    label_root_pth = "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/label/".format(file_list_n)
    #原图路径
    img_root_pth =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/img_o/".format(file_list_n)
    #分割结果
    seg_root_pth =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/seg/".format(file_list_n)

    out_root_pth =  "/root/Spine/Spine/Spine_segmentation/nnUnet_CTSpine1K/data/nnUNet_raw_data_base/nnUNet_raw_data/Task056_VerSe/{}/out_ct_begin_27_-1_20231228_2/".format(file_list_n)



    name_list = []
    id1 = []
    id2 = []
    id3 = []
    # id4 = []

    for cell in sheet['A']:
        name_list.append(cell.value)
    for cell in sheet['B']:
        id1.append(cell.value)
    for cell in sheet['C']:
        id2.append(cell.value)
    for cell in sheet['D']:
        id3.append(cell.value)
    # for cell in sheet['E']:
    #     id4.append(cell.value)
    print(name_list)

    for id,PAn in tqdm(enumerate(name_list)):
        # print("{},{}".format(id,PAn))

        excel_list_cur = []

        img = sitk.ReadImage(os.path.join(img_root_pth,"imagesTs_{}_{}/{}_0000.nii.gz".format(sheet_name,PAn,PAn)))
        img = sitk.GetArrayFromImage(img)
        c,h,w = img.shape

        label = sitk.ReadImage(os.path.join(label_root_pth,sheet_name,"{}.nii.gz".format(PAn)))
        label = sitk.GetArrayFromImage(label)

        seg = sitk.ReadImage(os.path.join(seg_root_pth,"predict_{}_{}/{}.nii.gz".format(sheet_name,PAn,PAn)))
        seg = sitk.GetArrayFromImage(seg)

        out = sitk.ReadImage(os.path.join(out_root_pth,"predict_{}_{}/{}{}_K{}C{}k{}i{}/view_all_corte.nii.gz".format(sheet_name,PAn,sheet_name,PAn,str(K_seg),str(C_seg),str(k_dili),str(iter_dili))))
        out = sitk.GetArrayFromImage(out)

        
        ##得到mask-3d 松质的
        mask_3d = np.zeros((c,h,w))
        for spine_id in spine_id_l:
            mask_3d[label==spine_id] = img[label==spine_id]
        mask_seg = np.zeros((c,h,w))
        for spine_id in seg_id_l:
            mask_seg[seg==spine_id] = img[seg==spine_id]
        
      
    #计算四个面的密度

        #获得切片位置
        z1 = id1[id]
        z2 = id2[id]
        z3 = id3[id]
        # z4 = id4[id]

        s1_img = mask_3d[z1,:,:]
        intensity_1 = np.sum(s1_img)
        counter_this_slice_1 = np.count_nonzero(s1_img)

        s2_img = mask_3d[z2,:,:]
        intensity_2 = np.sum(s2_img)
        counter_this_slice_2 = np.count_nonzero(s2_img)

        s3_img = mask_3d[z3,:,:]
        intensity_3 = np.sum(s3_img)
        counter_this_slice_3 = np.count_nonzero(s3_img)

        # s4_img = mask_3d[z4,:,:]
        # intensity_4 = np.sum(s4_img)
        # counter_this_slice_4 = np.count_nonzero(s4_img)

        #计算本患者的抽样平均密度
        # density_s = (intensity_1+intensity_2+intensity_3+intensity_4)/(counter_this_slice_1+counter_this_slice_2+counter_this_slice_3+counter_this_slice_4)

        density_s = (intensity_1+intensity_2+intensity_3)/(counter_this_slice_1+counter_this_slice_2+counter_this_slice_3)


        #计算标注的整体
        intensity_label = np.sum(mask_3d)
        counter_this_slice_label = np.count_nonzero(mask_3d,axis=(0,1,2))
        density_label = intensity_label/counter_this_slice_label

        #计算抽样和标注的差异
        result_1 = (density_s-density_label)/density_label
        ans_1.append(result_1)


        #令seg中皮质mask部分为0
        mask_seg[out==1] = 0
        intensity_seg = np.sum(mask_seg)
        counter_this_slice_seg = np.count_nonzero(mask_seg,axis=(0,1,2))
        density_seg = intensity_seg/counter_this_slice_seg

        #计算分割和标注的差异
        result_2 = (density_seg-density_label)/density_label
        ans_2.append(result_2)

        excel_list_cur.append(sheet_name+"_"+PAn)
        excel_list_cur.append(result_1)
        excel_list_cur.append(result_2)
        excel_list.append(excel_list_cur)

        
        # print("{} {} result1: {} ; result2: {} ".format(sheet,PAn,result_1,result_2))


#
#写入exel
write_to_excel(excel_path,sheet_str,info,excel_list)
ans_1 = np.array(ans_1)
ans_2 = np.array(ans_2)
np.save('ans_1_193.npy',ans_1) 
np.save('ans_2_193.npy',ans_2)
#绘制曲线图

# ave1 = sum(ans_1)/len(ans_1)
# ave2 = sum(ans_2)/len(ans_2)
# ave_err = (ave2-ave1)/ave1
# print("ave1 {}; ave2 {} ;ave_err {}".format(ave1,ave2,ave_err))

# plt.xlabel('Serial Number of Valid Image Dataset (n=192)')  # x轴标题
# plt.ylabel('Differences in Bone Mineral Density')  # y轴标题
# num = len(ans_1)
# x = [i for i in range(1,num+1)]
# # plt.plot(x,ans_1,marker='o',label =" sampling group")
# # plt.plot(x,ans_2,marker='x',label ="machine learning group")
# plt.scatter(x,ans_1,marker='o',label =" sampling group")
# plt.scatter(x,ans_2,marker='x',label ="machine learning group")
# # plt.legend(["select","seg"])
# plt.axhline(0,ls = '-', c = 'r',label ="0")
# plt.axhline(ave1,ls = '-', c = 'g',label ="mean sampling:{}".format(str(round(ave1, 3))))
# plt.axhline(ave2,ls = '-', c = 'b',label ="mean machine learning:{}".format(str(round(ave2, 3))))
# plt.legend(loc="upper right")
# f = plt.gcf()  #获取当前图像
# f.savefig(r'select_Gall_0126_2.png')
# f.clear()  #释放内存



print("end")